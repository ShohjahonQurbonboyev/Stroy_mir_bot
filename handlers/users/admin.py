from aiogram import types
from data.config import ADMINS, PASSWORD_ADMINS, USERS_CHANNEL
from loader import dp, db, bot
import pandas as pd
from states.states import  admin
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import os
from datetime import time, datetime
from keyboards.default.default_keyboards import  back_markup, create_main_admin, delete_btn, data_btn, add_code
from aiogram.dispatcher.storage import FSMContext


@dp.message_handler(text="/admin", state="*", user_id=ADMINS)
async def enter_admin_page(message: types.Message): 
    await message.answer("Parolni tering 👇", reply_markup=back_markup)
    await admin.password.set()

@dp.message_handler(state=admin.password)
async def confirm_admin(message: types.Message):
    try:
        if str(message.text) == PASSWORD_ADMINS:
            await message.delete()
            await message.answer("Admin panelga xush kelibsiz ✅", reply_markup=create_main_admin())
            await admin.menu.set()
        else:
            await message.answer("Parol notog'ri ❌")
    except  Exception as ex:
        await message.answer(ex)
    


@dp.message_handler(text = "➕ Promokod", state=admin.menu)
async def create_promokod(message:types.Message):
    await message.answer("Qaysi turda qo'shmoqchisiz ?", reply_markup=add_code())
    await admin.change_type_code.set()



@dp.message_handler(text = "➕ Donalab qo'shish", state=admin.change_type_code)
async def create_promokod(message:types.Message):
    await message.answer("Promokodni kiriting 👇", reply_markup=back_markup)
    await admin.promokod.set()

@dp.message_handler(state=admin.promokod)
async def create_promokod(message:types.Message, state : FSMContext):
    try:
        promokod = str(message.text)
        code = await db.select_code(code=promokod)
        if code is None:
            await state.update_data(data={"promokod": promokod})
            await message.answer("Bu kodga necha ball bermoqchisiz ?", reply_markup=back_markup)
            await admin.bal.set()
        else:
            await message.reply("Bunday promokod allaqchon mavjud\nIltimos boshqa promokod kiriting 👇")
    except  Exception as ex:
        await message.answer(ex)


@dp.message_handler(state=admin.bal)
async def create_ball(message:types.Message, state : FSMContext):
    try:
        data = await state.get_data()
        promocode = data["promokod"]
        ball = message.text
        if ball.isnumeric():
            await db.add_code(
                code=promocode,
                point=str(ball),
                used=True
            )
            await message.answer(f"<i>'{promocode}'</i> promokodiga <i>'{ball}'</i> bali biriktirildi va bazaga saqlandi ✅")
            await message.answer("Yana promokod kiriting 👇")
            await admin.promokod.set()
        else:
            await message.answer("Ball faqat son korinishida bo'lishi kerak ❌")
    except  Exception as ex:
        await message.answer(ex)
    


@dp.message_handler(text = "🗂 Excelda qo'shish", state=admin.change_type_code)
async def handle_document(message: types.Message):
    await message.answer("Excel fileni tashlang ",reply_markup=back_markup)
    await admin.excel.set()



@dp.message_handler(content_types=types.ContentType.DOCUMENT, state=admin.excel)
async def handle_document(message: types.Message):
    try:
        # Foydalanuvchi tomonidan yuborilgan faylni yuklash
        document = message.document
        file_id = document.file_id
        file = await bot.get_file(file_id)
        downloaded_file = await bot.download_file(file.file_path)
            
        # Faylni joylashtirish
        with open('Test_data.xlsx', 'wb') as new_file:
            new_file.write(downloaded_file.read())
    
        
        # Faylni o'qish va ma'lumotlarni saqlash
        wb = load_workbook(filename='Test_data.xlsx')
        sheet = wb.active
        for info in sheet.iter_rows(min_row=1, max_row=5000, min_col=1, max_col=2, values_only=True):
            promokod = str(info[0])
            if promokod and info[1] is not None:
                if str(info[1]).isnumeric():
                    code = await db.select_code(code=promokod)
                    if code is None:
                        await db.add_code(
                            code=promokod,
                            point=str(info[1]),
                            used=True
                        )
                else:
                    await message.answer(f"Bu ball son ko'rinishida bo'lishi kerak {info[1]}")
        await bot.send_document(chat_id=USERS_CHANNEL, document=file_id)
        os.remove('Test_data.xlsx')
        await message.answer("Malumotlar qabul qilindi ✅",reply_markup= back_markup)
    except Exception as ex:
        await message.answer(ex)

    



@dp.message_handler(text = "🗑 O'chirish", state=admin.menu)
async def delete(message:types.Message):
    await message.answer("Qaysi ma'lumotni o'chirmoqchisiz ?", reply_markup=delete_btn())
    await admin.delete.set()


@dp.message_handler(text = "❌ Barcha userlar ❌", state=admin.delete)
async def delete(message:types.Message):
    try:
        await db.delete_users()
        await message.answer("Barcha userlar o'chirib tashlandi ✅", reply_markup=delete_btn())
    except  Exception as ex:
        await message.answer(ex)


@dp.message_handler(text = "❌ User ❌", state=admin.delete)
async def delete(message:types.Message):
    await message.answer("Userning telegram Id sini kiriting 👇", reply_markup=back_markup)
    await admin.delete_from_id.set()
    

@dp.message_handler(state=admin.delete_from_id)
async def delete(message:types.Message):
    try:
        if message.text:
            id = int(message.text)
            user = await db.select_user(telegram_id=id)
            if user is None:
                await message.answer("Bunday id dagi foydalanuvchi topilmadi ❌")
            else:
                await db.delete_user(telegram_id=id)
                await message.answer("Foydalanuvchi muvofaqiyatli o'chirib tashlandi ✅")
        else:
            await message.answer("Telegram id faqat son ko'rinishida  kiritilishi kerak ❌")
    except  Exception as ex:
        await message.answer(ex)


@dp.message_handler(text = "❌ Barcha promokodlar ❌", state=admin.delete)
async def delete(message:types.Message):
    try:
        await db.delete_codes()
        await message.answer("Barcha promokodlar o'chirib tashlandi ✅", reply_markup=delete_btn())
    except  Exception as ex:
        await message.answer(ex)



@dp.message_handler(text = "📁 Malumotlar", state=admin.menu)
async def delete(message:types.Message):
    await message.reply("Qaysi turdagi ma'lumotni tekshirmoqchisiz ?", reply_markup=data_btn())
    await admin.data.set()


@dp.message_handler(text = "🚻 Foydalanuvchilar", state=admin.data)
async def data(message:types.Message):
    try:
        users = await db.select_all_users()
        id = []
        name = []
        for user in users:
            id.append(user[-1])
            name.append(user[1])
        data = {
            "Telegram ID": id,
            "Name": name
        }
        pd.options.display.max_rows = 10000
        df = pd.DataFrame(data)
        if len(df) > 50:
            for x in range(0, len(df), 50):
                await bot.send_message(message.chat.id, df[x:x + 50])
        else:
            await bot.send_message(message.chat.id, df)
    except  Exception as ex:
        await message.answer(ex)
    


@dp.message_handler(text = "👥 Foydalanuvchilar soni", state=admin.data)
async def delete(message:types.Message, state : FSMContext):
    count = await db.count_users()
    await  message.reply(f"Foydalanuvchilar soni {count} ta 😁")


@dp.message_handler(text = "👤 Foydalanuvchi haqida", state=admin.data)
async def delete(message:types.Message, state : FSMContext):
    await  message.reply(f"Kerakli userning telegram ID sini kiriting 👇", reply_markup=back_markup)
    await  admin.user_id.set()


@dp.message_handler(state=admin.user_id)
async def delete(message:types.Message, state : FSMContext):
    try:
        if message.text.isnumeric():
            users = await db.select_user(telegram_id=int(message.text))
            if users is None:
                await message.answer("Bunday id dagi foydalanuvchi mavjud emas !")
            else:
                if users[5] is None:
                    username = "Mavjud emas"
                else:
                    username = f"@{users[5]}"
                await message.answer_photo(photo=users[4], caption=f"Foydalanuvchi ma'lumotlari:\n\nIsmi : {users[1]}\n\nFamiliyasi : {users[2]}\n\nTel.raqam : {users[3]}\n\nUsername : {username}\n\nBall : {users[6]}\n\nTelegram_id : {users[7]}")
        else:
            await message.answer("Id faqat son ko'rinishida bo'lishi kerak !")
    except  Exception as ex:
        await message.answer(ex)
    





@dp.message_handler(text = "✏️ O'zgartirish", state=admin.menu)
async def delete(message:types.Message, state : FSMContext):
    await message.reply("Foydalanuvchining telegram Id sini kiriting 👇", reply_markup=back_markup)
    await admin.update_bal.set()



@dp.message_handler(state=admin.update_bal)
async def delete(message:types.Message, state : FSMContext):
    try:
        if message.text.isnumeric():
            users = await db.select_user(telegram_id=int(message.text))
            if users is None:
                await message.answer("Bunday id dagi foydalanuvchi mavjud emas !")
            else:
                await state.update_data(data={"user_id" : message.text})
                await message.answer(f"Bu Id {users[1]} ga tegishli uning balini necha qilib belgilamaoqchisiz ?", reply_markup=back_markup)
                await admin.update.set()
        else:
            await message.answer("Telegram id faqat son ko'rinishida bo'lishi kerak !")
    except  Exception as ex:
        await message.answer(ex)



@dp.message_handler(state=admin.update)
async def delete(message:types.Message, state : FSMContext):
    try:
        data = await state.get_data()
        if message.text.isnumeric():
            if int(message.text) >= 0:
                telegram_id  = int(data["user_id"])
                await  db.update_user_acoount(
                    account=str(message.text),
                    telegram_id=telegram_id
                    )
                user = await db.select_user(telegram_id=telegram_id)
                await message.answer(f"O'zgartirish muvaffaqiyatli yakunlandi✅\n\n{user[1]} {user[2]} ning bali endi {user[6]} ga teng ✅")
            else:
                await message.answer("Siz 0 dan kicik son kirita olmaysiz")
        else:
            await message.answer("Siz ballni faqat son ko'rinishida kirita olasiz !")
    except  Exception as ex:
        await message.answer(ex)